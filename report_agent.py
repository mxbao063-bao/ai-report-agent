#!/usr/bin/env python3
"""Generate a compact business report from CSV or optional XLSX input."""

from __future__ import annotations

import argparse
import csv
from collections import Counter, defaultdict
from pathlib import Path
from statistics import mean
from typing import Iterable


def parse_number(value: str) -> float | None:
    cleaned = value.strip().replace(",", "").replace("$", "")
    if cleaned.endswith("%"):
        cleaned = cleaned[:-1]
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def load_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        rows = [{key: (value or "").strip() for key, value in row.items()} for row in reader]
        return list(reader.fieldnames or []), rows


def load_xlsx(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("Excel input requires openpyxl: python3 -m pip install openpyxl") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(cell or "").strip() for cell in rows[0]]
    records: list[dict[str, str]] = []
    for values in rows[1:]:
        records.append({headers[index]: str(value or "").strip() for index, value in enumerate(values)})
    return headers, records


def load_data(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return load_xlsx(path)
    return load_csv(path)


def infer_numeric_columns(headers: Iterable[str], rows: list[dict[str, str]]) -> list[str]:
    numeric: list[str] = []
    for header in headers:
        values = [row.get(header, "") for row in rows if row.get(header, "").strip()]
        if not values:
            continue
        parsed = [parse_number(value) for value in values]
        success = sum(value is not None for value in parsed)
        if success / len(values) >= 0.7:
            numeric.append(header)
    return numeric


def infer_categorical_columns(headers: Iterable[str], rows: list[dict[str, str]], numeric_columns: set[str]) -> list[str]:
    categorical: list[str] = []
    for header in headers:
        if header in numeric_columns:
            continue
        values = [row.get(header, "").strip() for row in rows if row.get(header, "").strip()]
        unique = set(values)
        if values and 1 < len(unique) <= min(12, max(3, len(values) // 2 + 1)):
            categorical.append(header)
    return categorical


def number_values(rows: list[dict[str, str]], column: str) -> list[float]:
    values: list[float] = []
    for row in rows:
        value = parse_number(row.get(column, ""))
        if value is not None:
            values.append(value)
    return values


def format_number(value: float) -> str:
    if abs(value) >= 100:
        return f"{value:,.0f}"
    return f"{value:,.2f}".rstrip("0").rstrip(".")


def build_report(headers: list[str], rows: list[dict[str, str]], source_name: str) -> str:
    numeric_columns = infer_numeric_columns(headers, rows)
    categorical_columns = infer_categorical_columns(headers, rows, set(numeric_columns))

    lines = [
        f"# Business Report: {source_name}",
        "",
        "## Executive Summary",
        "",
        f"- Rows analyzed: {len(rows)}",
        f"- Columns analyzed: {len(headers)}",
        f"- Numeric metrics found: {', '.join(numeric_columns) if numeric_columns else 'none'}",
        f"- Segment fields found: {', '.join(categorical_columns) if categorical_columns else 'none'}",
        "",
    ]

    if numeric_columns:
        lines.extend(["## Metric Summary", "", "| Metric | Total | Average | Min | Max |", "| --- | ---: | ---: | ---: | ---: |"])
        for column in numeric_columns:
            values = number_values(rows, column)
            if values:
                lines.append(
                    f"| {column} | {format_number(sum(values))} | {format_number(mean(values))} | "
                    f"{format_number(min(values))} | {format_number(max(values))} |"
                )
        lines.append("")

    if categorical_columns and numeric_columns:
        segment = categorical_columns[0]
        metric = numeric_columns[0]
        grouped: defaultdict[str, float] = defaultdict(float)
        for row in rows:
            value = parse_number(row.get(metric, ""))
            if value is not None:
                grouped[row.get(segment, "Unknown") or "Unknown"] += value
        lines.extend([f"## Top Segments by {metric}", "", f"| {segment} | {metric} total |", "| --- | ---: |"])
        for name, total in sorted(grouped.items(), key=lambda item: item[1], reverse=True)[:10]:
            lines.append(f"| {name} | {format_number(total)} |")
        lines.append("")

    missing = Counter()
    for row in rows:
        for header in headers:
            if not row.get(header, "").strip():
                missing[header] += 1
    lines.extend(["## Data Quality", ""])
    if missing:
        for column, count in missing.most_common():
            lines.append(f"- `{column}` has {count} missing value(s).")
    else:
        lines.append("- No missing values detected.")

    lines.extend(
        [
            "",
            "## Suggested Next Actions",
            "",
            "- Confirm which metric is the primary business KPI.",
            "- Add a recurring export step so this report can run weekly.",
            "- Compare this period against the previous period once historical data is available.",
            "",
        ]
    )
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate a business report from CSV/XLSX data.")
    parser.add_argument("input", type=Path, help="Path to a CSV or XLSX file.")
    parser.add_argument("--output", type=Path, default=Path("report.md"), help="Markdown output path.")
    args = parser.parse_args()

    headers, rows = load_data(args.input)
    report = build_report(headers, rows, args.input.name)
    args.output.write_text(report, encoding="utf-8")
    print(f"Wrote {args.output}")


if __name__ == "__main__":
    main()
